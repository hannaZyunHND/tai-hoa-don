from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .renderer import TCHAT_LABELS, parse_invoice_date

BANG_KE_HEADERS = [
    "Mẫu số HD",
    "Ký hiệu hóa  đơn",
    "Số hóa đơn",
    "Ngày lập hóa đơn",
    "Ngày người bán ký số",
    "MCCQT",
    "Ngày CQT ký số",
    "Đơn vị tiền tệ",
    "Tỷ giá",
    "Tên người bán",
    "MST người bán",
    "Địa chỉ người bán",
    "Tên người mua",
    "MST người mua",
    "Địa chỉ người mua",
    "Mã VT",
    "Tên hàng hóa, dịch vụ",
    "Đơn vị tính",
    "Số lượng",
    "Đơn giá",
    "Chiết khấu",
    "Thuế suất",
    "Thành tiền chưa thuế",
    "Tiền thuế",
    "Tổng tiền CKTM",
    "Tổng tiền phí",
    "Tổng tiền thanh toán",
    "Trạng thái hóa đơn",
    "Kết quả kiểm tra hóa đơn",
    "url  tra cứu hóa đơn",
    "Mã tra cứu",
    "Ghi chú 1",
    "Hình  thức thanh toán",
    "Tính chất",
    "Ghi chú 2",
    "Số lô ",
    "Hạn dùng ",
]

TTHAI_LABELS = {
    1: "Hóa đơn mới",
    2: "Hóa đơn thay thế",
    3: "Hóa đơn điều chỉnh",
    4: "Hóa đơn đã bị thay thế",
    5: "Hóa đơn đã bị điều chỉnh",
    6: "Hóa đơn hủy",
}

TTTBAO_LABELS = {
    0: "Không mã",
    1: "Đã cấp mã hóa đơn",
    2: "Từ chối cấp mã",
    3: "Đã cấp mã (hóa đơn máy tính tiền)",
}

HTTTOAN_LABELS = {
    1: "Tiền mặt",
    2: "Chuyển khoản",
    3: "TM/CK",
    4: "Đối trừ công nợ",
    5: "Không thu tiền",
    6: "Khác",
    7: "Thẻ",
    8: "Ví điện tử",
    9: "TM/CK",
}

# ---- Theme ----
COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_TITLE_BG = "0F3A5F"
COLOR_ACCENT = "2E75B6"
COLOR_ZEBRA = "F2F7FB"
COLOR_PARAM_KEY = "E8F1F8"
COLOR_BORDER = "BDD7EE"
COLOR_LINK = "0563C1"
COLOR_MONEY = "FFF2CC"
COLOR_OK = "C6EFCE"

THIN = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)
HEADER_FILL = PatternFill("solid", fgColor=COLOR_HEADER_BG)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=COLOR_HEADER_FG)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=COLOR_HEADER_FG)
LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
BODY_FONT = Font(name="Calibri", size=11, color="1A1A1A")
LINK_FONT = Font(name="Calibri", size=11, color=COLOR_LINK, underline="single", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _fmt_date(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    dt = parse_invoice_date(text)
    if dt:
        return dt.strftime("%d/%m/%Y")
    try:
        raw = text.replace("Z", "+00:00") if text.endswith("Z") else text
        parsed = datetime.fromisoformat(raw)
        return parsed.strftime("%d/%m/%Y")
    except ValueError:
        return text


def _cks_signing_time(raw: Any) -> str:
    if not raw:
        return ""
    data = raw
    if isinstance(raw, str):
        try:
            data = json.loads(raw)
        except (TypeError, json.JSONDecodeError):
            return ""
    if isinstance(data, dict):
        return _fmt_date(data.get("SigningTime"))
    return ""


def _parse_tax_rate(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text.endswith("%"):
        try:
            return float(text[:-1].replace(",", ".")) / 100.0
        except ValueError:
            return text
    upper = text.upper()
    if upper in {"KKKNT", "KCT", "KHONGCHIU", "KHÔNG CHỊU"}:
        return text
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return text


def _field_from_ttkhac(items: Any, *names: str) -> str:
    if not items:
        return ""
    wanted = {n.lower() for n in names}
    for item in items:
        if not isinstance(item, dict):
            continue
        key = str(item.get("ttruong") or "").lower()
        if key in wanted:
            return str(item.get("dlieu") or "")
    return ""


def _relative_link_path(path: Path, base_dir: Path) -> str:
    """Path tương đối so với thư mục chứa Excel (dùng / để portable khi zip)."""
    try:
        rel = path.resolve().relative_to(base_dir.resolve())
    except ValueError:
        rel = Path(path.name)
    return rel.as_posix()


def _set_file_hyperlink(
    cell,
    path: Path | None,
    label: str,
    *,
    base_dir: Path,
) -> None:
    """Hyperlink relative path — zip gửi đi click vẫn mở được."""
    if path is None or not path.exists():
        cell.value = ""
        return
    rel = _relative_link_path(path, base_dir)
    # Escape dấu " trong path (hiếm) để công thức HYPERLINK an toàn
    safe_rel = rel.replace('"', '""')
    cell.value = f'=HYPERLINK("{safe_rel}","{label}")'
    cell.font = LINK_FONT
    cell.alignment = CENTER


def _style_header_row(ws, row: int, col_count: int, *, height: float = 32) -> None:
    ws.row_dimensions[row].height = height
    for col in range(1, col_count + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN


def _style_data_cell(cell, *, zebra: bool = False, money: bool = False, center: bool = False) -> None:
    cell.font = BODY_FONT
    cell.border = THIN
    if money:
        cell.alignment = RIGHT
        cell.number_format = "#,##0"
        cell.fill = PatternFill("solid", fgColor=COLOR_MONEY if not zebra else "FCE4A8")
    elif center:
        cell.alignment = CENTER
        if zebra:
            cell.fill = PatternFill("solid", fgColor=COLOR_ZEBRA)
    else:
        cell.alignment = LEFT
        if zebra:
            cell.fill = PatternFill("solid", fgColor=COLOR_ZEBRA)


def _autosize_columns(ws, widths: dict[str, float] | list[float], col_count: int) -> None:
    if isinstance(widths, dict):
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width
        return
    for i, width in enumerate(widths, start=1):
        if i <= col_count:
            ws.column_dimensions[get_column_letter(i)].width = width


def load_bang_ke_headers(template_path: Path) -> list[str]:
    if not template_path.exists():
        return list(BANG_KE_HEADERS)
    wb = load_workbook(template_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers: list[str] = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value) if cell.value is not None else "")
        return headers or list(BANG_KE_HEADERS)
    finally:
        wb.close()


def build_bang_ke_rows(detail: dict[str, Any]) -> list[list[Any]]:
    lines = detail.get("hdhhdvu") or []
    if not lines:
        lines = [{}]

    base_common = {
        "mau": detail.get("khmshdon"),
        "khhdon": detail.get("khhdon"),
        "shdon": detail.get("shdon"),
        "ngay_lap": _fmt_date(detail.get("tdlap")),
        "ngay_ky_nb": _cks_signing_time(detail.get("nbcks")) or _fmt_date(detail.get("nky")),
        "mccqt": detail.get("mhdon") or "",
        "ngay_ky_cqt": _cks_signing_time(detail.get("cqtcks")) or _fmt_date(detail.get("ncma")),
        "dvtte": detail.get("dvtte") or "",
        "tgia": detail.get("tgia") if detail.get("tgia") is not None else "",
        "nbten": detail.get("nbten") or "",
        "nbmst": detail.get("nbmst") or "",
        "nbdchi": detail.get("nbdchi") or "",
        "nmten": detail.get("nmten") or "",
        "nmmst": detail.get("nmmst") or "",
        "nmdchi": detail.get("nmdchi") or "",
        "ttcktmai": detail.get("ttcktmai") if detail.get("ttcktmai") is not None else "",
        "tgtphi": detail.get("tgtphi") if detail.get("tgtphi") is not None else "",
        "tgtttbso": detail.get("tgtttbso") if detail.get("tgtttbso") is not None else "",
        "tthai": TTHAI_LABELS.get(detail.get("tthai"), detail.get("tthai") or ""),
        "tttbao": TTTBAO_LABELS.get(detail.get("tttbao"), detail.get("tttbao") or ""),
        "url": _field_from_ttkhac(detail.get("ttkhac"), "url", "Url", "URL", "Link")
        or _field_from_ttkhac(detail.get("nbttkhac"), "url", "Url", "URL"),
        "ma_tra_cuu": detail.get("mhso")
        or _field_from_ttkhac(detail.get("ttkhac"), "KeySearch", "Fkey", "MaTraCuu")
        or "",
        "gchu": detail.get("gchu") or "",
        "htttoan": detail.get("thtttoan")
        or HTTTOAN_LABELS.get(detail.get("htttoan"), detail.get("htttoan") or ""),
    }

    rows: list[list[Any]] = []
    for line in lines:
        ma_vt = line.get("mhhdvu") or line.get("mau") or line.get("ma") or ""
        thue_suat = _parse_tax_rate(
            line.get("ltsuat") if line.get("ltsuat") is not None else line.get("tsuat")
        )
        tchat = TCHAT_LABELS.get(line.get("tchat"), line.get("tchat") or "")
        rows.append(
            [
                base_common["mau"],
                base_common["khhdon"],
                base_common["shdon"],
                base_common["ngay_lap"],
                base_common["ngay_ky_nb"],
                base_common["mccqt"],
                base_common["ngay_ky_cqt"],
                base_common["dvtte"],
                base_common["tgia"],
                base_common["nbten"],
                base_common["nbmst"],
                base_common["nbdchi"],
                base_common["nmten"],
                base_common["nmmst"],
                base_common["nmdchi"],
                ma_vt,
                line.get("ten") or "",
                line.get("dvtinh") or "",
                line.get("sluong") if line.get("sluong") is not None else "",
                line.get("dgia") if line.get("dgia") is not None else "",
                line.get("stckhau") if line.get("stckhau") is not None else "",
                thue_suat,
                line.get("thtien") if line.get("thtien") is not None else "",
                line.get("tthue") if line.get("tthue") is not None else "",
                base_common["ttcktmai"],
                base_common["tgtphi"],
                base_common["tgtttbso"],
                base_common["tthai"],
                base_common["tttbao"],
                base_common["url"],
                base_common["ma_tra_cuu"],
                base_common["gchu"],
                base_common["htttoan"],
                tchat,
                "",
                "",
                "",
            ]
        )
    return rows


def write_export_workbook(
    *,
    output_path: Path,
    params: list[tuple[str, Any]],
    invoice_rows: list[dict[str, Any]],
    bang_ke_rows: list[list[Any]],
    template_path: Path,
) -> None:
    headers = load_bang_ke_headers(template_path)
    wb = Workbook()
    base_dir = output_path.parent

    # ========== Sheet 1: Tham số tải ==========
    ws1 = wb.active
    ws1.title = "Tham so tai"
    ws1.merge_cells("A1:B1")
    title = ws1["A1"]
    title.value = "THAM SỐ LƯỢT TẢI HOÁ ĐƠN"
    title.font = TITLE_FONT
    title.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title.alignment = CENTER
    ws1.row_dimensions[1].height = 36

    ws1["A2"] = "Tham số"
    ws1["B2"] = "Giá trị"
    _style_header_row(ws1, 2, 2, height=24)

    for idx, (key, value) in enumerate(params, start=3):
        key_cell = ws1.cell(idx, 1, key)
        val_cell = ws1.cell(idx, 2, value if value is not None else "")
        key_cell.font = LABEL_FONT
        key_cell.fill = PatternFill("solid", fgColor=COLOR_PARAM_KEY)
        key_cell.border = THIN
        key_cell.alignment = LEFT
        val_cell.font = BODY_FONT
        val_cell.border = THIN
        val_cell.alignment = LEFT
        if key in {"Số hoá đơn tải thành công"} and isinstance(value, int) and value > 0:
            val_cell.fill = PatternFill("solid", fgColor=COLOR_OK)
        ws1.row_dimensions[idx].height = 20

    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 78
    ws1.freeze_panes = "A3"

    # ========== Sheet 2: Danh sách hoá đơn ==========
    ws2 = wb.create_sheet("Danh sach hoa don")
    ws2.merge_cells("A1:L1")
    t2 = ws2["A1"]
    t2.value = "DANH SÁCH HOÁ ĐƠN ĐÃ TẢI"
    t2.font = TITLE_FONT
    t2.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    t2.alignment = CENTER
    ws2.row_dimensions[1].height = 36

    list_headers = [
        "STT",
        "Mẫu số",
        "Ký hiệu",
        "Số hóa đơn",
        "Ngày lập",
        "MST người bán",
        "Tên người bán",
        "Tổng tiền thanh toán",
        "Trạng thái",
        "XML",
        "HTML",
        "PDF",
    ]
    for col, title_text in enumerate(list_headers, start=1):
        ws2.cell(2, col, title_text)
    _style_header_row(ws2, 2, len(list_headers), height=28)

    money_cols = {8}
    center_cols = {1, 2, 3, 4, 5, 6, 9, 10, 11, 12}

    for row_idx, item in enumerate(invoice_rows, start=3):
        zebra = (row_idx % 2) == 1
        values = [
            item.get("stt"),
            item.get("khmshdon"),
            item.get("khhdon"),
            item.get("shdon"),
            item.get("ngay_lap"),
            item.get("nbmst"),
            item.get("nbten"),
            item.get("tgtttbso"),
            item.get("tthai"),
            None,
            None,
            None,
        ]
        for col, value in enumerate(values, start=1):
            if col >= 10:
                continue
            cell = ws2.cell(row_idx, col, value)
            _style_data_cell(
                cell,
                zebra=zebra,
                money=col in money_cols,
                center=col in center_cols,
            )
        _set_file_hyperlink(
            ws2.cell(row_idx, 10), item.get("xml_path"), "xml", base_dir=base_dir
        )
        _set_file_hyperlink(
            ws2.cell(row_idx, 11), item.get("html_path"), "html", base_dir=base_dir
        )
        _set_file_hyperlink(
            ws2.cell(row_idx, 12), item.get("pdf_path"), "pdf", base_dir=base_dir
        )
        for col in (10, 11, 12):
            cell = ws2.cell(row_idx, col)
            cell.border = THIN
            if zebra:
                cell.fill = PatternFill("solid", fgColor=COLOR_ZEBRA)
        ws2.row_dimensions[row_idx].height = 20

    _autosize_columns(
        ws2,
        [6, 10, 12, 12, 12, 16, 42, 18, 16, 8, 8, 8],
        len(list_headers),
    )
    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:L{max(2, 2 + len(invoice_rows))}"

    # ========== Sheet 3: Bảng kê chi tiết ==========
    ws3 = wb.create_sheet("Bang ke chi tiet")
    last_col = get_column_letter(len(headers))
    ws3.merge_cells(f"A1:{last_col}1")
    t3 = ws3["A1"]
    t3.value = "BẢNG KÊ CHI TIẾT HÀNG HOÁ / DỊCH VỤ"
    t3.font = TITLE_FONT
    t3.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    t3.alignment = CENTER
    ws3.row_dimensions[1].height = 36

    for col, title_text in enumerate(headers, start=1):
        ws3.cell(2, col, title_text)
    _style_header_row(ws3, 2, len(headers), height=40)

    # Cột tiền trong bảng kê (1-based): 20 Đơn giá, 21 CK, 23-27 tiền
    bang_ke_money_cols = {19, 20, 21, 23, 24, 25, 26, 27}
    bang_ke_center_cols = {1, 2, 3, 4, 5, 7, 8, 9, 18, 22, 28, 33, 34}

    for row_idx, row in enumerate(bang_ke_rows, start=3):
        zebra = (row_idx % 2) == 1
        for col_idx, value in enumerate(row, start=1):
            if col_idx > len(headers):
                break
            cell = ws3.cell(row_idx, col_idx, value)
            _style_data_cell(
                cell,
                zebra=zebra,
                money=col_idx in bang_ke_money_cols and isinstance(value, (int, float)),
                center=col_idx in bang_ke_center_cols,
            )
            # Thuế suất dạng 0.08 hiện %
            if col_idx == 22 and isinstance(value, float) and value <= 1:
                cell.number_format = "0%"
                cell.alignment = CENTER
        ws3.row_dimensions[row_idx].height = 18

    default_widths = {
        "A": 10,
        "B": 12,
        "C": 12,
        "D": 13,
        "E": 14,
        "F": 22,
        "G": 13,
        "H": 10,
        "I": 8,
        "J": 28,
        "K": 15,
        "L": 28,
        "M": 24,
        "N": 14,
        "O": 28,
        "P": 12,
        "Q": 36,
        "R": 10,
        "S": 10,
        "T": 12,
        "U": 10,
        "V": 10,
        "W": 14,
        "X": 12,
        "Y": 12,
        "Z": 12,
        "AA": 14,
        "AB": 14,
        "AC": 18,
        "AD": 18,
        "AE": 14,
        "AF": 12,
        "AG": 14,
        "AH": 14,
        "AI": 10,
        "AJ": 10,
        "AK": 10,
    }
    _autosize_columns(ws3, default_widths, len(headers))
    ws3.freeze_panes = "A3"
    if bang_ke_rows:
        ws3.auto_filter.ref = f"A2:{last_col}{2 + len(bang_ke_rows)}"

    # Tab màu
    ws1.sheet_properties.tabColor = "1F4E79"
    ws2.sheet_properties.tabColor = "2E75B6"
    ws3.sheet_properties.tabColor = "548235"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
